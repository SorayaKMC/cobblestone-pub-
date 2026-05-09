"""Cobblestone Pub Management App."""

from functools import wraps
from flask import Flask, redirect, url_for, request, Response, render_template, jsonify
from datetime import date
import os
import re
import secrets
import threading

# Force-load googleapiclient on the MAIN thread before any background
# thread spawns. This works around a known thread-safety bug where two
# threads first-importing googleapiclient.discovery concurrently can
# return a partially-initialised module:
#   https://github.com/googleapis/google-api-python-client/issues/1502
# The downstream modules (drive_watcher, gmail_poller, payroll_drafts,
# calendar_client, tips_sheet_importer) all import these at module top,
# but they themselves are only imported inside threads — so the race
# still happens on first-import. Importing here forces the cache hit.
try:
    from google.oauth2 import service_account as _sa  # noqa: F401
    from googleapiclient.discovery import build as _build  # noqa: F401
    from googleapiclient.http import (  # noqa: F401
        MediaIoBaseDownload as _mdl,
        MediaIoBaseUpload as _mul,
    )
except ImportError as e:
    # Local dev without the deps installed — log and continue. On Render
    # the deps are always present so this won't fire in production.
    print(f"[boot] googleapiclient not installed: {e}")

import db
import config


def pdf_basename(path):
    """Return a human-friendly filename from a stored pdf_path.

    Strips the directory and the `YYYYMMDD_HHMMSS_` upload-timestamp prefix
    that save_uploaded_pdf() prepends, so the original uploaded name is shown.
    """
    if not path:
        return ""
    name = os.path.basename(str(path))
    # Drop leading timestamp if present: 20260417_113245_Original.pdf -> Original.pdf
    return re.sub(r"^\d{8}_\d{6}_", "", name)


def eu_date(value):
    """Format a date for display in European convention: dd/mm/yyyy.

    Storage stays in ISO (YYYY-MM-DD) for sortability — this is the
    display-only conversion. Accepts ISO strings, datetime/date objects,
    or pre-formatted strings (returned as-is if already dd/mm/yyyy).

    Empty / None / unparseable values become an empty string so templates
    don't render literal 'None'.
    """
    from datetime import datetime as _dt, date as _date
    if value in (None, "", "None"):
        return ""
    # Already a datetime/date
    if isinstance(value, _dt):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, _date):
        return value.strftime("%d/%m/%Y")
    s = str(value).strip()
    # ISO date YYYY-MM-DD (with optional time component)
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if iso_match:
        y, mo, d = iso_match.group(1), iso_match.group(2), iso_match.group(3)
        return f"{d}/{mo}/{y}"
    # Already dd/mm/yyyy — pass through
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        return s
    return s


def pretty_date(value):
    """Format a date as 'DD-Mmm-YYYY' (e.g. '05-May-2026') for the bookings list."""
    from datetime import datetime as _dt, date as _date
    if value in (None, "", "None"):
        return ""
    if isinstance(value, _dt):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, _date):
        return value.strftime("%d-%b-%Y")
    s = str(value).strip()
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if iso_match:
        try:
            d = _date(int(iso_match.group(1)), int(iso_match.group(2)), int(iso_match.group(3)))
            return d.strftime("%d-%b-%Y")
        except ValueError:
            pass
    return s


def day_name(value):
    """Return the day-of-week name (e.g. 'Tuesday') for an ISO date string or date object."""
    from datetime import datetime as _dt, date as _date
    if value in (None, "", "None"):
        return ""
    if isinstance(value, _dt):
        return value.strftime("%A")
    if isinstance(value, _date):
        return value.strftime("%A")
    s = str(value).strip()
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if iso_match:
        try:
            d = _date(int(iso_match.group(1)), int(iso_match.group(2)), int(iso_match.group(3)))
            return d.strftime("%A")
        except ValueError:
            pass
    return ""


def eu_month(value):
    """Format an ISO date as 'mmm yyyy' (e.g. 'Apr 2026'). For headers
    and summary blocks where a full date is unnecessary."""
    from datetime import datetime as _dt
    if value in (None, "", "None"):
        return ""
    s = str(value).strip()
    iso_match = re.match(r"^(\d{4})-(\d{2})", s)
    if iso_match:
        try:
            return _dt(int(iso_match.group(1)), int(iso_match.group(2)), 1).strftime("%b %Y")
        except Exception:
            return s
    return s


def _gmail_poll_loop():
    """Background thread: poll invoice@cobblestonepub.ie every 30 minutes.

    Disabled silently if GOOGLE_SERVICE_ACCOUNT_JSON is not set, so local
    development works without any Google credentials.
    """
    import time
    if not config.GOOGLE_SERVICE_ACCOUNT_JSON:
        print("[gmail] GOOGLE_SERVICE_ACCOUNT_JSON not set — inbox polling disabled")
        return
    print(f"[gmail] Inbox polling active (every {config.GMAIL_POLL_INTERVAL}s)")
    while True:
        try:
            from gmail_poller import check_inbox
            results = check_inbox()
            saved = sum(1 for r in results if r.get("invoice_id"))
            if saved:
                print(f"[gmail] Pulled {saved} new invoice(s) from inbox")
            elif results:
                print(f"[gmail] Checked inbox — nothing new")
        except Exception as e:
            print(f"[gmail] Poll error: {e}")
        time.sleep(config.GMAIL_POLL_INTERVAL)


def _pto_weekly_recalc_loop():
    """Background thread: weekly PTO recalculation, Sunday 23:00 Dublin time.

    Recalculates the past 4 ISO weeks for every employee so any backdated
    Square timecard edits get picked up. Side-effect: refreshes the 13-week
    avg shift used in the accrual formula. Idempotent — safe if it runs
    twice within the same window.
    """
    import time
    from datetime import datetime, timedelta
    try:
        from zoneinfo import ZoneInfo
        DUBLIN = ZoneInfo("Europe/Dublin")
    except Exception:
        DUBLIN = None

    def _now_dublin():
        return datetime.now(DUBLIN) if DUBLIN else datetime.now()

    def _most_recent_sunday_2300(now):
        """Most recent Sunday 23:00 that has already occurred."""
        days_back = (now.weekday() - 6) % 7  # Mon=0 .. Sun=6
        sun = (now - timedelta(days=days_back)).replace(
            hour=23, minute=0, second=0, microsecond=0
        )
        if sun > now:
            sun -= timedelta(days=7)
        return sun

    print("[pto-weekly] Auto-recalc thread active (checks every hour for Sunday 23:00 Dublin)")
    while True:
        try:
            now = _now_dublin()
            trigger = _most_recent_sunday_2300(now)
            last, _ = db.get_cache("pto_weekly_recalc_last_run")
            should_run = True
            if last and "ts" in last:
                try:
                    last_ts = datetime.fromisoformat(last["ts"])
                    if last_ts.tzinfo is None and DUBLIN:
                        last_ts = last_ts.replace(tzinfo=DUBLIN)
                    if last_ts >= trigger:
                        should_run = False
                except Exception:
                    pass

            if should_run:
                print(f"[pto-weekly] Triggering recalc (trigger={trigger.isoformat()}, now={now.isoformat()})")
                from_date = (now - timedelta(weeks=4)).date().isoformat()
                to_date = now.date().isoformat()
                try:
                    import pto_engine
                    import square_client
                    team_members = square_client.get_team_members()
                    categories = db.get_employee_categories()
                    count = 0
                    for cat in categories:
                        try:
                            pto_engine.recalculate_pto(
                                cat["team_member_id"], from_date, to_date, team_members
                            )
                            count += 1
                        except Exception as e:
                            print(f"[pto-weekly] {cat['family_name']}: {e}")
                    db.set_cache("pto_weekly_recalc_last_run", {
                        "ts": now.isoformat(),
                        "from": from_date, "to": to_date,
                        "employees": count,
                    })
                    print(f"[pto-weekly] Recalculated {count} employees ({from_date} to {to_date})")
                except Exception as e:
                    print(f"[pto-weekly] Recalc failed: {e}")
        except Exception as e:
            print(f"[pto-weekly] Loop error: {e}")
        time.sleep(3600)  # check every hour


def _drive_watch_loop():
    """Background thread: scan the invoices Drive folder every 30 minutes."""
    import time
    if not config.GOOGLE_SERVICE_ACCOUNT_JSON:
        print("[drive] GOOGLE_SERVICE_ACCOUNT_JSON not set — Drive watcher disabled")
        return
    if not config.GOOGLE_DRIVE_INVOICES_FOLDER_ID:
        print("[drive] GOOGLE_DRIVE_INVOICES_FOLDER_ID not set — Drive watcher disabled")
        return
    # Stagger first run so we don't fight the gmail poller at boot
    time.sleep(60)
    print(f"[drive] Folder watcher active (every {config.GMAIL_POLL_INTERVAL}s)")
    while True:
        try:
            from drive_watcher import import_pending
            results = import_pending()
            imported = sum(1 for r in results if r.get("invoice_id"))
            errored = sum(1 for r in results if r.get("error"))
            if imported or errored:
                print(f"[drive] Imported {imported}, errors: {errored}, "
                      f"total scanned: {len(results)}")
            db.set_cache("drive_watcher_last_run", {
                "ts": __import__("datetime").datetime.now().isoformat(),
                "imported": imported,
                "errored": errored,
                "total": len(results),
                "results": results[-10:],  # keep last 10 for display
            })
        except Exception as e:
            print(f"[drive] Scan error: {e}")
            db.set_cache("drive_watcher_last_run", {
                "ts": __import__("datetime").datetime.now().isoformat(),
                "error": str(e),
            })
        time.sleep(config.GMAIL_POLL_INTERVAL)


def _warmup_cache():
    """Pre-populate dashboard cache in a background thread at startup.

    Designed for low memory: fetches one week at a time, writes to SQLite,
    then forces garbage collection before moving to the next week.
    Stays well under 100MB of peak memory.
    """
    try:
        import time
        import gc
        time.sleep(10)  # let app finish booting & handle first requests
        from routes.dashboard import _get_week_sales_with_daily, _get_week_payroll
        import square_client

        current_year, current_week = square_client.current_week()
        print(f"[warmup] Starting cache warmup for {current_year} W02-W{current_week}")

        for week in range(2, current_week + 1):
            try:
                # Fetch + cache current year
                _get_week_sales_with_daily(current_year, week)
                gc.collect()
                # Fetch + cache previous year (for YoY)
                _get_week_sales_with_daily(current_year - 1, week)
                gc.collect()
                # Fetch + cache payroll
                _get_week_payroll(current_year, week)
                gc.collect()
                print(f"[warmup] W{week:02d} cached")
                time.sleep(1)  # be gentle on Square's rate limits
            except Exception as e:
                print(f"[warmup] W{week:02d} failed: {e}")

        print("[warmup] Complete")
    except Exception as e:
        print(f"[warmup] Thread failed: {e}")


def check_auth(username, password):
    """Constant-time comparison of provided creds against configured ones."""
    if not config.AUTH_ENABLED:
        return True
    return (
        secrets.compare_digest(username or "", config.AUTH_USERNAME)
        and secrets.compare_digest(password or "", config.AUTH_PASSWORD)
    )


def require_auth(f):
    """Decorator to protect routes with HTTP Basic Auth."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not config.AUTH_ENABLED:
            return f(*args, **kwargs)
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return Response(
                "Authentication required.",
                401,
                {"WWW-Authenticate": 'Basic realm="Cobblestone Pub Manager"'},
            )
        return f(*args, **kwargs)
    return wrapper


def create_app():
    app = Flask(__name__)
    app.secret_key = os.getenv("FLASK_SECRET_KEY", "cobblestone-pub-local-app-2026")
    app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB file upload limit

    @app.context_processor
    def inject_globals():
        return {"today": date.today().strftime("%A, %d %B %Y")}

    app.jinja_env.filters["pdf_basename"] = pdf_basename
    app.jinja_env.filters["eu_date"] = eu_date
    app.jinja_env.filters["eu_month"] = eu_month
    app.jinja_env.filters["pretty_date"] = pretty_date
    app.jinja_env.filters["day_name"] = day_name

    # Initialize database
    with app.app_context():
        db.init_db()

    # Background warmup - populates dashboard cache so it doesn't time out
    if os.getenv("ENABLE_WARMUP", "1") == "1":
        threading.Thread(target=_warmup_cache, daemon=True).start()

    # Background Gmail polling - checks invoice inbox every 30 minutes
    threading.Thread(target=_gmail_poll_loop, daemon=True).start()

    # Background Drive watcher - scans the invoices folder for human-uploaded PDFs
    threading.Thread(target=_drive_watch_loop, daemon=True).start()

    # Background weekly PTO recalc - runs every Sunday 23:00 Dublin time
    threading.Thread(target=_pto_weekly_recalc_loop, daemon=True).start()

    # Apply HTTP Basic Auth globally (if enabled via env vars)
    if config.AUTH_ENABLED:
        @app.before_request
        def protect_all():
            # Allow static files and health checks without auth
            if request.endpoint in ("static", "healthz"):
                return None
            # Public booking form + band portal + Square webhooks (no auth required)
            if request.path.startswith("/book") or request.path.startswith("/webhooks"):
                return None
            # Cron endpoints — protected by their own key check, not Basic Auth
            if request.path == "/admin/run-reminders":
                return None
            auth = request.authorization
            if not auth or not check_auth(auth.username, auth.password):
                return Response(
                    "Please log in to access the Cobblestone Pub Manager.",
                    401,
                    {"WWW-Authenticate": 'Basic realm="Cobblestone Pub Manager"'},
                )

    # Register blueprints
    from routes import settings, payroll, dashboard, pto, bookkeeping, bookings
    app.register_blueprint(settings.bp)
    app.register_blueprint(payroll.bp)
    app.register_blueprint(dashboard.bp)
    app.register_blueprint(pto.bp)
    app.register_blueprint(bookkeeping.bp)
    app.register_blueprint(bookings.bp)

    @app.route("/")
    def index():
        from datetime import datetime
        from routes.dashboard import _get_week_sales_with_daily
        import square_client as sq

        week_net = None
        week_label = ""
        week_dates = ""
        try:
            current_year, current_week = sq.current_week()
            sales = _get_week_sales_with_daily(current_year, current_week)
            if sales:
                week_net = round(sales["total"])
            week_label = f"W{current_week:02d}"
            start, end = sq.week_dates(current_year, current_week)
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d")
            week_dates = f"{start_dt.strftime('%-d %b')} – {end_dt.strftime('%-d %b')}"
        except Exception:
            pass

        hour = datetime.now().hour
        if hour < 12:
            greeting = "Good morning"
        elif hour < 17:
            greeting = "Good afternoon"
        else:
            greeting = "Good evening"

        return render_template("home.html",
            greeting=greeting,
            week_net=week_net,
            week_label=week_label,
            week_dates=week_dates,
        )

    @app.route("/healthz")
    def healthz():
        return "ok", 200

    # ── One-time admin import endpoint ────────────────────────────────────
    # Protected by ADMIN_PASSWORD. Upload the xlsx, runs the historical
    # import, shows results. Safe to leave in place — locked behind password.
    @app.route("/admin/import-bookings", methods=["GET", "POST"])
    def admin_import_bookings():
        import tempfile, os as _os

        # Simple password gate (POST param or query string)
        provided = (request.form.get("password") or
                    request.args.get("password") or "")
        authed = secrets.compare_digest(provided, config.ADMIN_PASSWORD)

        if request.method == "GET":
            # Show login form (no password submitted yet)
            return render_template("admin_import.html",
                                   authed=False, results=None, error=None)

        if not authed:
            return render_template("admin_import.html",
                                   authed=False, results=None,
                                   error="Wrong password — try again.")

        # Password correct — check for file
        f = request.files.get("xlsx")
        if not f or not f.filename:
            return render_template("admin_import.html",
                                   authed=True, results=None,
                                   error="No file selected.")

        if not f.filename.lower().endswith(".xlsx"):
            return render_template("admin_import.html",
                                   authed=True, results=None,
                                   error="Please upload a .xlsx file.")

        # Save to a temp file and run the import
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                f.save(tmp.name)
                tmp_path = tmp.name

            from openpyxl import load_workbook
            import bookings_historical_import as bhi

            db.init_db()
            wb = load_workbook(tmp_path, data_only=True)
            results = {}

            if "Bookings" in wb.sheetnames:
                ins, skip, err = bhi.import_main_sheet(wb["Bookings"], dry_run=False)
                results["Bookings"] = {"inserted": ins, "skipped": skip, "errors": err}

            archive_name = next((n for n in wb.sheetnames if n.startswith("Archive")), None)
            if archive_name:
                ins, skip, err = bhi.import_archive_sheet(wb[archive_name], dry_run=False)
                results[archive_name] = {"inserted": ins, "skipped": skip, "errors": err}

            counts = db.booking_counts()
            _os.unlink(tmp_path)

            return render_template("admin_import.html",
                                   authed=True, results=results,
                                   counts=counts, error=None)
        except Exception as e:
            return render_template("admin_import.html",
                                   authed=True, results=None,
                                   error=f"Import failed: {e}")

    # ── Daily reminder cron endpoint ─────────────────────────────────────
    # Call this daily (e.g. Render Cron Job) to send 3-day reminder emails.
    # URL: /admin/run-reminders?key=<CRON_KEY>
    # CRON_KEY env var defaults to ADMIN_PASSWORD if not set separately.
    @app.route("/admin/run-reminders")
    def run_reminders():
        from datetime import timedelta
        import bookings_email

        cron_key = os.getenv("CRON_KEY", config.ADMIN_PASSWORD)
        provided  = request.args.get("key", "")
        if not secrets.compare_digest(provided, cron_key):
            return jsonify({"error": "Forbidden"}), 403

        target_date = (date.today() + timedelta(days=3)).isoformat()
        bookings    = db.list_bookings(status="confirmed",
                                       start_date=target_date,
                                       end_date=target_date)

        sent = skipped = errors = 0
        base_url = request.host_url.rstrip("/")
        today_iso = date.today().isoformat()

        for b in bookings:
            # Skip if reminder already logged today (prevents double-send)
            conn = db.get_db()
            already = conn.execute(
                """SELECT 1 FROM booking_audit
                   WHERE booking_id=? AND action='reminder_sent'
                     AND created_at >= ?""",
                (b["id"], today_iso),
            ).fetchone()
            conn.close()

            if already:
                skipped += 1
                continue

            try:
                ok = bookings_email.send_booking_reminder(b, base_url)
                if ok:
                    db.add_booking_audit(b["id"], "system", "reminder_sent",
                                         f"3-day reminder → {b['contact_email']}")
                    sent += 1
                else:
                    skipped += 1   # SMTP not configured or no email address
            except Exception as e:
                db.add_booking_audit(b["id"], "system", "reminder_error", str(e))
                errors += 1

        # ── Auto-complete past bookings ───────────────────────────────────
        completed_count = db.auto_complete_past_bookings()

        # ── Door-person alert (7-day window) ─────────────────────────────
        door_pending = db.get_bookings_needing_door_confirmation(days_ahead=7)
        door_alert_sent = False
        if door_pending:
            # Only send once per day — check audit on the first booking in the list
            conn = db.get_db()
            already_alerted = conn.execute(
                """SELECT 1 FROM booking_audit
                   WHERE action='door_alert_sent' AND created_at >= ?""",
                (today_iso,),
            ).fetchone()
            conn.close()

            if not already_alerted:
                try:
                    ok = bookings_email.send_door_person_alert(door_pending, base_url)
                    if ok:
                        # Log against the first booking as a proxy record
                        ids = ", ".join(str(b["id"]) for b in door_pending)
                        db.add_booking_audit(
                            door_pending[0]["id"], "system", "door_alert_sent",
                            f"Door person alert sent for booking IDs: {ids}",
                        )
                        door_alert_sent = True
                except Exception as e:
                    print(f"[reminders] Door person alert error: {e}")

        return jsonify({
            "target_date":          target_date,
            "bookings_found":       len(bookings),
            "reminders_sent":       sent,
            "skipped":              skipped,
            "errors":               errors,
            "auto_completed":       completed_count,
            "door_unconfirmed":     len(door_pending),
            "door_alert_sent":      door_alert_sent,
        })

    # ── Square webhook ───────────────────────────────────────────────────────
    # Receives payment.updated events from Square. When a door fee payment
    # completes, auto-stamps door_fee_paid_at on the matching booking.
    # Register this URL in the Square Developer dashboard:
    #   https://cobblestone-pub.onrender.com/webhooks/square
    # Add the resulting signature key as SQUARE_WEBHOOK_SIGNATURE_KEY in Render.
    @app.route("/webhooks/square", methods=["POST"])
    def square_webhook():
        import hmac as _hmac
        import hashlib
        import base64

        payload = request.get_data()
        sig_header = request.headers.get("x-square-hmacsha256-signature", "")
        sig_key = config.SQUARE_WEBHOOK_SIGNATURE_KEY

        # Verify signature when a key is configured
        if sig_key:
            combined = request.url.encode("utf-8") + payload
            computed = base64.b64encode(
                _hmac.new(sig_key.encode("utf-8"), combined, hashlib.sha256).digest()
            ).decode("utf-8")
            if not _hmac.compare_digest(computed, sig_header):
                print("[webhook/square] Signature mismatch — rejected")
                return jsonify({"error": "Invalid signature"}), 403

        data = request.get_json(force=True, silent=True) or {}
        event_type = data.get("type", "")

        # We only care about completed payments
        if event_type == "payment.updated":
            payment = data.get("data", {}).get("object", {}).get("payment", {})
            if payment.get("status") != "COMPLETED":
                return jsonify({"ok": True}), 200

            note = payment.get("note", "")
            booking_id = None
            if "cobblestone_booking_id:" in note:
                try:
                    booking_id = int(note.split("cobblestone_booking_id:")[1].strip())
                except (ValueError, IndexError):
                    pass

            if booking_id:
                booking = db.get_booking(booking_id)
                if booking and not booking["door_fee_paid_at"]:
                    now_iso = date.today().isoformat()
                    conn = db.get_db()
                    conn.execute(
                        "UPDATE bookings SET door_fee_paid_at=?, updated_at=? WHERE id=?",
                        (now_iso, now_iso, booking_id),
                    )
                    conn.commit()
                    conn.close()
                    payment_id = payment.get("id", "")
                    db.add_booking_audit(
                        booking_id, "system", "door_fee_paid",
                        f"€50 door fee paid online via Square · payment ID: {payment_id}",
                    )
                    print(f"[webhook/square] Door fee marked paid for booking #{booking_id}")

        return jsonify({"ok": True}), 200

    # ── SMTP test endpoint ───────────────────────────────────────────────────
    # Hit this to verify email is working without going through a full booking.
    # URL: /admin/test-email?key=<ADMIN_PASSWORD>&to=you@example.com
    @app.route("/admin/test-email")
    def test_email():
        provided = request.args.get("key", "")
        if not secrets.compare_digest(provided, config.ADMIN_PASSWORD):
            return jsonify({"error": "Forbidden"}), 403

        to = request.args.get("to", "")
        if not to:
            return jsonify({"error": "Pass ?to=your@email.com"}), 400

        # Report current SMTP config (passwords masked)
        smtp_status = {
            "SMTP_HOST":     config.SMTP_HOST or "(not set)",
            "SMTP_PORT":     config.SMTP_PORT,
            "SMTP_USERNAME": config.SMTP_USERNAME or "(not set)",
            "SMTP_PASSWORD": "****" if config.SMTP_PASSWORD else "(not set)",
            "BOOKING_FROM":  config.BOOKING_FROM or "(not set)",
        }

        import smtplib
        try:
            with smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT, timeout=10) as server:
                server.ehlo()
                server.starttls()
                server.login(config.SMTP_USERNAME, config.SMTP_PASSWORD)
                server.sendmail(
                    config.BOOKING_FROM or config.SMTP_USERNAME,
                    to,
                    f"Subject: Cobblestone SMTP test\r\n\r\nSMTP is working correctly.",
                )
            return jsonify({"status": "sent", "to": to, "config": smtp_status})
        except Exception as e:
            return jsonify({"status": "failed", "error": str(e), "config": smtp_status}), 500

    return app


app = create_app()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, port=port)
