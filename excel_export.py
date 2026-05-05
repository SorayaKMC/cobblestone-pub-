"""Excel export for Cobblestone Pub payroll reports.

Generates two formats:
1. "For Peter" - formatted payroll for the accountant
2. "Raw Timecards" - Square timecard data
"""

from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
HEADER_FILL = PatternFill("solid", fgColor="343A40")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
MONEY_FORMAT = '#,##0.00'
HOURS_FORMAT = '0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
TOTAL_FILL = PatternFill("solid", fgColor="E9ECEF")
TOTAL_FONT = Font(bold=True, size=10, name="Arial")
BODY_FONT = Font(size=10, name="Arial")


def _apply_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def _apply_body(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def generate_peter_excel(week_label, payroll_data, net_sales=None):
    """Generate the 'for Peter' payroll Excel.

    Columns:
      A=(blank) B=First C=Last D=Wage E=Gross F=Hours G=Tips H=Cleaning
      I=Bonus J=Holiday Hrs K=Holiday Pay L=Total M=(gap) N=Total for labor
      O=Upper Management P=Management Q=Staff R=Staff+M
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{week_label} for Peter"

    headers = [
        "", "First", "Last", "Wage", "Gross", "Hours", "Tips",
        "Cleaning", "Bonus", "Holiday Hrs", "Holiday Pay", "Total",
        "", "Total for labor",
        "Upper Management", "Management", "Staff", "Staff+M"
    ]

    for col, h in enumerate(headers, 1):
        _apply_header(ws, 1, col, h)

    widths = [4, 14, 18, 8, 10, 8, 8, 10, 10, 11, 11, 10, 2, 14, 18, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = w

    row = 2
    um_total = Decimal("0")
    mgmt_total = Decimal("0")
    staff_total = Decimal("0")

    for emp in payroll_data:
        wage = Decimal(str(emp["wage_rate"]))
        holiday_hrs = Decimal(str(emp.get("holiday_hours", 0) or 0))
        holiday_pay = (holiday_hrs * wage).quantize(Decimal("0.01"))

        _apply_body(ws, row, 2, emp["given_name"])
        _apply_body(ws, row, 3, emp["family_name"])
        _apply_body(ws, row, 4, float(wage), MONEY_FORMAT)
        _apply_body(ws, row, 5, float(emp["gross"]), MONEY_FORMAT)
        _apply_body(ws, row, 6, float(emp["hours"]), HOURS_FORMAT)
        _apply_body(ws, row, 7, float(emp["tips"]), MONEY_FORMAT)
        _apply_body(ws, row, 8, float(emp["cleaning"]), MONEY_FORMAT)
        _apply_body(ws, row, 9, float(emp.get("bonus", 0)), MONEY_FORMAT)
        _apply_body(ws, row, 10, float(holiday_hrs), HOURS_FORMAT)
        _apply_body(ws, row, 11, float(holiday_pay), MONEY_FORMAT)
        _apply_body(ws, row, 12, float(emp["total"]), MONEY_FORMAT)

        # Column M is the gap, N onwards is Total for labor + categories
        _apply_body(ws, row, 13, emp["category"].split()[0] if emp["category"] == "Upper Management" else emp["category"])
        _apply_body(ws, row, 14, float(emp["total_for_labor"]), MONEY_FORMAT)

        cat = emp["category"]
        if cat == "Upper Management":
            _apply_body(ws, row, 15, float(emp["total_for_labor"]), MONEY_FORMAT)
            um_total += emp["total_for_labor"]
        elif cat == "Management":
            _apply_body(ws, row, 16, float(emp["total_for_labor"]), MONEY_FORMAT)
            mgmt_total += emp["total_for_labor"]
        elif cat == "Staff":
            _apply_body(ws, row, 17, float(emp["total_for_labor"]), MONEY_FORMAT)
            staff_total += emp["total_for_labor"]

        row += 1

    # Totals row
    total_row = row
    for col in range(1, 19):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER

    ws.cell(row=total_row, column=2, value="TOTALS").font = TOTAL_FONT

    if len(payroll_data) > 0:
        data_start = 2
        data_end = total_row - 1
        ws.cell(row=total_row, column=5,  value=f"=SUM(E{data_start}:E{data_end})").number_format = MONEY_FORMAT
        ws.cell(row=total_row, column=6,  value=f"=SUM(F{data_start}:F{data_end})").number_format = HOURS_FORMAT
        ws.cell(row=total_row, column=7,  value=f"=SUM(G{data_start}:G{data_end})").number_format = MONEY_FORMAT
        ws.cell(row=total_row, column=8,  value=f"=SUM(H{data_start}:H{data_end})").number_format = MONEY_FORMAT
        ws.cell(row=total_row, column=9,  value=f"=SUM(I{data_start}:I{data_end})").number_format = MONEY_FORMAT
        ws.cell(row=total_row, column=10, value=f"=SUM(J{data_start}:J{data_end})").number_format = HOURS_FORMAT
        ws.cell(row=total_row, column=11, value=f"=SUM(K{data_start}:K{data_end})").number_format = MONEY_FORMAT
        ws.cell(row=total_row, column=12, value=f"=SUM(L{data_start}:L{data_end})").number_format = MONEY_FORMAT
        ws.cell(row=total_row, column=14, value=f"=SUM(N{data_start}:N{data_end})").number_format = MONEY_FORMAT

    # Category totals
    ws.cell(row=total_row, column=15, value=float(um_total)).number_format = MONEY_FORMAT
    ws.cell(row=total_row, column=16, value=float(mgmt_total)).number_format = MONEY_FORMAT
    ws.cell(row=total_row, column=17, value=float(staff_total)).number_format = MONEY_FORMAT
    ws.cell(row=total_row, column=18, value=float(mgmt_total + staff_total)).number_format = MONEY_FORMAT

    # Summary block below totals
    row = total_row + 2
    if net_sales is not None:
        ws.cell(row=row, column=2, value="Net Sales").font = TOTAL_FONT
        ws.cell(row=row, column=5, value=float(net_sales)).number_format = MONEY_FORMAT
        row += 1
        ws.cell(row=row, column=2, value="Total Labor").font = TOTAL_FONT
        ws.cell(row=row, column=5, value=float(um_total + mgmt_total + staff_total)).number_format = MONEY_FORMAT
        row += 1
        labor_total = um_total + mgmt_total + staff_total
        if net_sales > 0:
            labor_pct = float(labor_total / net_sales * 100)
            ws.cell(row=row, column=2, value="Labor %").font = TOTAL_FONT
            cell = ws.cell(row=row, column=5, value=labor_pct / 100)
            cell.number_format = '0.0%'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_raw_timecard_excel(week_label, timecard_data):
    """Generate raw timecard Excel matching Square export format.

    Args:
        week_label: e.g. "Week 15"
        timecard_data: list of dicts with keys:
            employee_id, given_name, family_name,
            regular_hours, overtime_hours, doubletime_hours, total_hours,
            regular_cost, overtime_cost, doubletime_cost, total_cost,
            transaction_tips, declared_cash_tips

    Returns BytesIO with .xlsx content.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = week_label

    headers = [
        "Employee number", "First name", "Last name",
        "Regular hours", "Overtime hours", "Doubletime hours", "Total paid hours",
        "Regular labor cost", "Overtime labor cost", "Doubletime labor cost",
        "Total labor cost", "Transaction tips", "Declared cash tips"
    ]

    for col, h in enumerate(headers, 1):
        _apply_header(ws, 1, col, h)

    widths = [16, 12, 18, 14, 14, 16, 16, 18, 18, 20, 16, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = w

    row = 2
    for emp in timecard_data:
        _apply_body(ws, row, 1, emp.get("employee_id", ""))
        _apply_body(ws, row, 2, emp["given_name"])
        _apply_body(ws, row, 3, emp["family_name"])
        _apply_body(ws, row, 4, float(emp["regular_hours"]), HOURS_FORMAT)
        _apply_body(ws, row, 5, float(emp["overtime_hours"]), HOURS_FORMAT)
        _apply_body(ws, row, 6, float(emp["doubletime_hours"]), HOURS_FORMAT)
        _apply_body(ws, row, 7, float(emp["total_hours"]), HOURS_FORMAT)

        # Format costs as EUR strings to match existing spreadsheets
        for col_idx, key in [(8, "regular_cost"), (9, "overtime_cost"),
                             (10, "doubletime_cost"), (11, "total_cost"),
                             (12, "transaction_tips"), (13, "declared_cash_tips")]:
            val = emp.get(key, Decimal("0"))
            _apply_body(ws, row, col_idx, f"EUR{float(val):.2f}")

        row += 1

    # Totals
    total_row = row
    for col in range(1, 14):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER

    ws.cell(row=total_row, column=2, value="TOTALS").font = TOTAL_FONT

    if len(timecard_data) > 0:
        for col in [4, 5, 6, 7]:
            ws.cell(row=total_row, column=col, value=f"=SUM({chr(64+col)}2:{chr(64+col)}{total_row-1})").number_format = HOURS_FORMAT

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_invoice_monthly_excel(period_label, invoices):
    """Monthly invoice summary for the accountant / VAT prep.

    Two sheets:
      'Invoices'  — one row per approved invoice with all fields
      'Summary'   — totals by category, by VAT rate, grand total

    Args:
        period_label: human-readable period (e.g. 'April 2026' or
            'April-May 2026' for a VAT period).
        invoices: list of approved invoice rows from db.list_invoices().

    Returns BytesIO of an .xlsx workbook.
    """
    wb = Workbook()

    # ─── Sheet 1: line items ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Invoices"

    title_cell = ws.cell(row=1, column=1,
                         value=f"Cobblestone Pub - Invoice Summary - {period_label}")
    title_cell.font = Font(bold=True, size=14, name="Arial")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    sub_cell = ws.cell(row=2, column=1,
                       value=f"{len(invoices)} approved invoice(s)")
    sub_cell.font = Font(italic=True, size=10, name="Arial", color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["Date", "Supplier", "Invoice #", "Category",
               "Net", "VAT %", "VAT", "Total", "Status"]
    for col, h in enumerate(headers, start=1):
        _apply_header(ws, row=4, col=col, value=h)

    row = 5
    total_net = total_vat = total_total = Decimal("0")
    for inv in sorted(invoices, key=lambda i: (i["invoice_date"] or "", i["supplier_name"] or "")):
        _apply_body(ws, row, 1, inv["invoice_date"] or "")
        _apply_body(ws, row, 2, inv["supplier_name"] or "")
        _apply_body(ws, row, 3, inv["invoice_number"] or "")
        _apply_body(ws, row, 4, inv["category"] or "")
        _apply_body(ws, row, 5, float(inv["net_amount"] or 0), MONEY_FORMAT)
        _apply_body(ws, row, 6, float(inv["vat_rate"] or 0), '0.0"%"')
        _apply_body(ws, row, 7, float(inv["vat_amount"] or 0), MONEY_FORMAT)
        _apply_body(ws, row, 8, float(inv["total_amount"] or 0), MONEY_FORMAT)
        _apply_body(ws, row, 9, (inv["status"] or "").title())
        total_net   += Decimal(str(inv["net_amount"] or 0))
        total_vat   += Decimal(str(inv["vat_amount"] or 0))
        total_total += Decimal(str(inv["total_amount"] or 0))
        row += 1

    # Totals row
    total_row = row
    _apply_body(ws, total_row, 1, "TOTAL")
    ws.cell(row=total_row, column=1).font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    for c in range(2, 10):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
    ws.cell(row=total_row, column=5, value=float(total_net)).number_format = MONEY_FORMAT
    ws.cell(row=total_row, column=7, value=float(total_vat)).number_format = MONEY_FORMAT
    ws.cell(row=total_row, column=8, value=float(total_total)).number_format = MONEY_FORMAT

    widths = [12, 28, 14, 18, 12, 8, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    # ─── Sheet 2: summary ────────────────────────────────────────────────────
    sm = wb.create_sheet("Summary")

    sm.cell(row=1, column=1, value=f"Summary - {period_label}").font = Font(bold=True, size=14, name="Arial")
    sm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    sm.cell(row=3, column=1, value="By Category").font = Font(bold=True, size=11, name="Arial")
    for col, h in enumerate(["Category", "Net", "VAT", "Total"], start=1):
        _apply_header(sm, row=4, col=col, value=h)

    by_cat = {}
    for inv in invoices:
        c = inv["category"] or "(uncategorised)"
        if c not in by_cat:
            by_cat[c] = {"net": Decimal("0"), "vat": Decimal("0"), "total": Decimal("0")}
        by_cat[c]["net"]   += Decimal(str(inv["net_amount"] or 0))
        by_cat[c]["vat"]   += Decimal(str(inv["vat_amount"] or 0))
        by_cat[c]["total"] += Decimal(str(inv["total_amount"] or 0))

    r = 5
    for cat, vals in sorted(by_cat.items()):
        _apply_body(sm, r, 1, cat)
        _apply_body(sm, r, 2, float(vals["net"]),   MONEY_FORMAT)
        _apply_body(sm, r, 3, float(vals["vat"]),   MONEY_FORMAT)
        _apply_body(sm, r, 4, float(vals["total"]), MONEY_FORMAT)
        r += 1
    _apply_body(sm, r, 1, "TOTAL")
    sm.cell(row=r, column=1).font = TOTAL_FONT
    sm.cell(row=r, column=1).fill = TOTAL_FILL
    for c in range(2, 5):
        sm.cell(row=r, column=c).fill = TOTAL_FILL
        sm.cell(row=r, column=c).font = TOTAL_FONT
        sm.cell(row=r, column=c).border = THIN_BORDER
    sm.cell(row=r, column=2, value=float(total_net)).number_format = MONEY_FORMAT
    sm.cell(row=r, column=3, value=float(total_vat)).number_format = MONEY_FORMAT
    sm.cell(row=r, column=4, value=float(total_total)).number_format = MONEY_FORMAT
    cat_total_row = r

    sm.cell(row=cat_total_row + 3, column=1, value="By VAT Rate").font = Font(bold=True, size=11, name="Arial")
    for col, h in enumerate(["VAT Rate", "Net", "VAT", "Total"], start=1):
        _apply_header(sm, row=cat_total_row + 4, col=col, value=h)

    by_rate = {}
    for inv in invoices:
        rate = float(inv["vat_rate"] or 0)
        if rate not in by_rate:
            by_rate[rate] = {"net": Decimal("0"), "vat": Decimal("0"), "total": Decimal("0")}
        by_rate[rate]["net"]   += Decimal(str(inv["net_amount"] or 0))
        by_rate[rate]["vat"]   += Decimal(str(inv["vat_amount"] or 0))
        by_rate[rate]["total"] += Decimal(str(inv["total_amount"] or 0))

    r = cat_total_row + 5
    for rate, vals in sorted(by_rate.items(), reverse=True):
        _apply_body(sm, r, 1, f"{rate:g}%")
        _apply_body(sm, r, 2, float(vals["net"]),   MONEY_FORMAT)
        _apply_body(sm, r, 3, float(vals["vat"]),   MONEY_FORMAT)
        _apply_body(sm, r, 4, float(vals["total"]), MONEY_FORMAT)
        r += 1

    for i, w in enumerate([20, 14, 14, 14], start=1):
        sm.column_dimensions[sm.cell(row=4, column=i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
