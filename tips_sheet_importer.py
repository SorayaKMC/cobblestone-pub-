"""Import weekly tips from the shared Google Sheet.

The sheet has one tab per pay week. Each tab has rows per employee and
columns for daily tips (and possibly a Total). This module fetches the
sheet via the Drive API (export-as-xlsx — no extra Sheets API scope
needed beyond the existing 'drive' scope), parses out per-employee
totals, and matches names to our DB by fuzzy lookup.

Heuristics on tab + column structure are intentional — the user's sheet
format isn't fixed in stone, and I'd rather flash 'matched X of Y' than
break on a layout drift.
"""

import io
import json
import re
from datetime import datetime, timedelta

import config
import db


def _drive_service():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    sa_info = json.loads(config.GOOGLE_SERVICE_ACCOUNT_JSON)
    creds = (
        service_account.Credentials
        .from_service_account_info(
            sa_info, scopes=["https://www.googleapis.com/auth/drive"],
        )
        .with_subject("info@cobblestonepub.ie")
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def fetch_sheet_as_xlsx(sheet_id):
    """Export the Google Sheet as an .xlsx file (all tabs preserved)."""
    service = _drive_service()
    request = service.files().export_media(
        fileId=sheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    from googleapiclient.http import MediaIoBaseDownload
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tab name → ISO week matching
# ---------------------------------------------------------------------------

def _iso_week_dates(iso_week):
    """Return (monday_date, sunday_date) for an ISO week string '2026-W18'."""
    year, w = iso_week.split("-W")
    year, w = int(year), int(w)
    # ISO week: Monday is weekday() == 0
    jan4 = datetime(year, 1, 4)
    monday_w1 = jan4 - timedelta(days=jan4.weekday())
    monday = monday_w1 + timedelta(weeks=w - 1)
    sunday = monday + timedelta(days=6)
    return monday.date(), sunday.date()


def _match_tab_to_week(tab_name, iso_week):
    """Return True if this tab's name plausibly refers to the given ISO week.

    Tries multiple formats:
      - 'Week 18'      → match week number
      - 'W18'          → match week number
      - '18'           → match week number
      - 'May 1', '1/5', '01/05/2026', '01-05-26' → match by date in week
      - 'Apr 27 - May 3' → range match
    """
    if not tab_name:
        return False
    name = tab_name.strip()
    name_lower = name.lower()
    year, w_num = iso_week.split("-W")
    w_num_int = int(w_num)
    monday, sunday = _iso_week_dates(iso_week)

    # Week-number patterns
    for pat in (r"\bweek\s*0*(\d{1,2})\b", r"\bw0*(\d{1,2})\b", r"\b0*(\d{1,2})\b"):
        m = re.search(pat, name_lower)
        if m and int(m.group(1)) == w_num_int:
            return True

    # Numeric date patterns: DD/MM, DD-MM, DD.MM, DD/MM/YYYY etc.
    for m in re.finditer(r"\b(\d{1,2})[/\-\.](\d{1,2})(?:[/\-\.](\d{2,4}))?\b", name):
        d = int(m.group(1))
        mo = int(m.group(2))
        y = m.group(3)
        if y:
            y = int(y) + (2000 if int(y) < 100 else 0)
        else:
            y = monday.year
        try:
            d_obj = datetime(y, mo, d).date()
        except ValueError:
            try:
                # Try MM/DD/YYYY interpretation
                d_obj = datetime(y, d, mo).date()
            except ValueError:
                continue
        if monday <= d_obj <= sunday:
            return True

    # Month-name patterns: "May 1", "1 May", "May 1 - May 3"
    months = {
        "january": 1, "jan": 1, "february": 2, "feb": 2,
        "march": 3, "mar": 3, "april": 4, "apr": 4,
        "may": 5, "june": 6, "jun": 6, "july": 7, "jul": 7,
        "august": 8, "aug": 8, "september": 9, "sep": 9, "sept": 9,
        "october": 10, "oct": 10, "november": 11, "nov": 11,
        "december": 12, "dec": 12,
    }
    pat = r"(\w{3,9})\s+(\d{1,2})"
    for m in re.finditer(pat, name_lower):
        mo = months.get(m.group(1))
        if not mo:
            continue
        d = int(m.group(2))
        try:
            d_obj = datetime(monday.year, mo, d).date()
        except ValueError:
            continue
        if monday <= d_obj <= sunday:
            return True
    pat = r"(\d{1,2})\s+(\w{3,9})"
    for m in re.finditer(pat, name_lower):
        mo = months.get(m.group(2))
        if not mo:
            continue
        d = int(m.group(1))
        try:
            d_obj = datetime(monday.year, mo, d).date()
        except ValueError:
            continue
        if monday <= d_obj <= sunday:
            return True

    return False


# ---------------------------------------------------------------------------
# Sheet structure parsing
# ---------------------------------------------------------------------------

def _is_text_cell(v):
    return isinstance(v, str) and v.strip()


def _is_numeric_cell(v):
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        s = v.strip().replace("€", "").replace(",", "").replace("$", "")
        if not s:
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False
    return False


def _to_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("€", "").replace(",", "").replace("$", "").replace("(", "-").replace(")", "")
        if not s:
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def _normalise_name(s):
    if not s:
        return ""
    return re.sub(r"[^\w]", " ", s).lower().strip()


def _name_tokens(s):
    return set(t for t in _normalise_name(s).split() if len(t) > 1)


# Rows whose name cell is one of these are footers, not employees
_FOOTER_NAMES = {"total", "totals", "grand total", "subtotal", "sub-total", "sub total", ""}

# Sheet uses nicknames or alt spellings that the fuzzy matcher can't reach.
# Map → canonical first name in our DB. Easy to extend; add new entries here.
SHEET_NAME_ALIASES = {
    "podge": "padhraig",  # Padhraig O'Maolagain
}


def _classify_row(row_name, active_employees, inactive_employees):
    """Classify a row by its name cell.

    Returns one of:
      ('footer',   None)              - row is a TOTAL / SUBTOTAL footer
      ('blank',    None)              - empty name
      ('active',   team_member_id)    - matches an active employee
      ('inactive', team_member_id)    - matches a former employee (skip silently)
      ('unknown',  None)              - matches no one we know
    """
    if not row_name:
        return ("blank", None)
    name_clean = str(row_name).strip().lower()
    if name_clean in _FOOTER_NAMES:
        return ("footer", None)

    raw_tokens = _name_tokens(row_name)
    if not raw_tokens:
        return ("blank", None)

    # Nickname aliases — replace alias token with canonical first-name token
    aliased_tokens = set()
    for tok in raw_tokens:
        aliased_tokens.add(SHEET_NAME_ALIASES.get(tok, tok))

    def _try_match(pool):
        # Strict: full token equality
        for emp in pool:
            emp_tokens = _name_tokens(f"{emp['given_name']} {emp['family_name']}")
            if aliased_tokens == emp_tokens:
                return emp["team_member_id"]
        # Looser: subset either way
        for emp in pool:
            emp_tokens = _name_tokens(f"{emp['given_name']} {emp['family_name']}")
            if aliased_tokens.issubset(emp_tokens) or emp_tokens.issubset(aliased_tokens):
                return emp["team_member_id"]
        # Loosest: first-name match (sheet uses given names only / alt spellings)
        for emp in pool:
            if emp["given_name"].lower() in aliased_tokens:
                return emp["team_member_id"]
        return None

    tm = _try_match(active_employees)
    if tm:
        return ("active", tm)
    tm = _try_match(inactive_employees)
    if tm:
        return ("inactive", tm)
    return ("unknown", None)


def _parse_sheet_for_week(ws, active_employees, inactive_employees=None):
    """Return {tm_id: total_tips_eur, '_unmatched': [(raw_name, value), ...]}.

    Heuristics:
      - Find the row that contains the most string cells > 2 chars (header row).
      - Find a 'Total' column by header text; otherwise fall back to:
        - The rightmost numeric column with non-zero values.
        - Or sum all numeric columns that aren't headers.
      - Names: first text-heavy column, looking for matches against active
        employees.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"_tabname": ws.title, "_unmatched": []}

    # --- Find header row ---
    header_row_idx = 0
    best_score = 0
    for i, row in enumerate(rows[:8]):  # check first 8 rows only
        score = sum(1 for c in row if _is_text_cell(c) and len(c) > 2)
        if score > best_score:
            best_score = score
            header_row_idx = i

    headers = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    headers_lower = [h.lower() for h in headers]

    # --- Find name column ---
    name_col = None
    for i, h in enumerate(headers_lower):
        if any(kw in h for kw in ("name", "employee", "staff", "person")):
            name_col = i
            break
    if name_col is None:
        # Fallback: leftmost column with text in body rows
        for col_idx in range(len(headers)):
            text_count = sum(
                1 for r in rows[header_row_idx + 1:header_row_idx + 12]
                if col_idx < len(r) and _is_text_cell(r[col_idx])
            )
            if text_count >= 3:
                name_col = col_idx
                break
    if name_col is None:
        return {"_tabname": ws.title, "_unmatched": [], "_error": "Could not identify name column"}

    # --- Find total column ---
    total_col = None
    for i, h in enumerate(headers_lower):
        if any(kw in h for kw in ("total", "sum", "weekly", "week total")):
            total_col = i
            break

    matched = {}
    unmatched = []
    skipped_inactive = []

    inactive_employees = inactive_employees or []

    for r in rows[header_row_idx + 1:]:
        if not r:
            continue
        if name_col >= len(r):
            continue
        raw_name = r[name_col]
        if not _is_text_cell(raw_name):
            continue
        if total_col is not None and total_col < len(r):
            value = _to_float(r[total_col])
        else:
            value = sum(
                _to_float(r[i])
                for i in range(len(r))
                if i != name_col and _is_numeric_cell(r[i])
            )
        if value <= 0:
            continue

        kind, tm_id = _classify_row(raw_name, active_employees, inactive_employees)
        if kind == "active":
            matched[tm_id] = matched.get(tm_id, 0.0) + value
        elif kind == "inactive":
            skipped_inactive.append((str(raw_name).strip(), round(value, 2)))
        elif kind == "unknown":
            unmatched.append((str(raw_name).strip(), round(value, 2)))
        # 'footer' / 'blank' silently dropped

    matched["_tabname"] = ws.title
    matched["_unmatched"] = unmatched
    matched["_skipped_inactive"] = skipped_inactive
    matched["_total_col_used"] = headers[total_col] if total_col is not None else "(summed daily columns)"
    matched["_name_col_used"] = headers[name_col] if name_col < len(headers) else "(column A)"
    return matched


def import_tips_for_week(iso_week):
    """Fetch the sheet, find the right tab for iso_week, save tips. Returns
    a dict suitable for flash-message display."""
    sheet_id = config.TIPS_SHEET_ID
    if not sheet_id:
        return {"ok": False, "error": "TIPS_SHEET_ID not configured"}
    if not config.GOOGLE_SERVICE_ACCOUNT_JSON:
        return {"ok": False, "error": "Service account not configured"}

    try:
        xlsx_bytes = fetch_sheet_as_xlsx(sheet_id)
    except Exception as e:
        return {"ok": False, "error": f"Could not fetch sheet: {e}"}

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"Could not parse spreadsheet: {e}"}

    # Split employees by active flag — active ones get tips applied,
    # inactive (former) ones are silently skipped to keep the unmatched
    # list focused on truly unknown names.
    all_emp_rows = db.get_employee_categories()
    active_employees = [
        {"team_member_id": r["team_member_id"],
         "given_name": r["given_name"],
         "family_name": r["family_name"]}
        for r in all_emp_rows if r["is_active"]
    ]
    inactive_employees = [
        {"team_member_id": r["team_member_id"],
         "given_name": r["given_name"],
         "family_name": r["family_name"]}
        for r in all_emp_rows if not r["is_active"]
    ]

    # Find the right tab
    matching_tab = None
    candidate_tabs = []
    for sheet_name in wb.sheetnames:
        candidate_tabs.append(sheet_name)
        if _match_tab_to_week(sheet_name, iso_week):
            matching_tab = sheet_name
            break

    if not matching_tab:
        return {
            "ok": False,
            "error": (f"No tab matched {iso_week}. Available tabs: "
                      f"{', '.join(candidate_tabs[:20])}"),
            "tabs_seen": candidate_tabs,
        }

    parsed = _parse_sheet_for_week(wb[matching_tab], active_employees, inactive_employees)
    if "_error" in parsed:
        return {"ok": False, "error": f"Tab '{matching_tab}': {parsed['_error']}"}

    tips_to_save = {
        k: v for k, v in parsed.items() if not k.startswith("_")
    }
    if tips_to_save:
        db.bulk_set_weekly_tips(iso_week, tips_to_save)

    return {
        "ok": True,
        "tab": matching_tab,
        "matched_count": len(tips_to_save),
        "matched_total": round(sum(tips_to_save.values()), 2),
        "unmatched": parsed.get("_unmatched", []),
        "skipped_inactive": parsed.get("_skipped_inactive", []),
        "name_col": parsed.get("_name_col_used"),
        "total_col": parsed.get("_total_col_used"),
    }
