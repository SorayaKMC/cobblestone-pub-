"""Import form responses (Cobblestone Backroom Event Questionnaire) into bookings DB.

Two-pass workflow:
  1. Dry run (default) — classifies each form submission and writes an
     import_review.md report. No DB writes.
  2. python3 import_form_responses.py <xlsx> --confirm — actually creates
     booking rows.

Classification:
  AUTO_IMPORT        form email + date matches a calendar entry exactly      → status=confirmed
  NEEDS_REVIEW       email matches calendar but date off by 1-3 days, OR    → status=inquiry
                     date matches but email differs                            (shows in bookings queue)
  PENDING            no calendar match at all                                 → status=inquiry
  PAST               event date is before today                               → skipped

Dedupe: when the same (email, event_date) appears more than once, the
latest submission (by Timestamp) wins.

Usage:
  python3 import_form_responses.py /path/to/responses.xlsx \
      --calendar-json /path/to/calendar_dump.json [--confirm]
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, date

from openpyxl import load_workbook

import db


# ─── Constants ──────────────────────────────────────────────────────────────
DATE_TOLERANCE_DAYS = 3
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")

# Form column header → schema field mapping
COL_MAP = {
    "Timestamp":                                                                                  "_timestamp",
    "Event date?\nHere is a link to the dates available for the backroom. Please confirm with us that the date IS available.": "_event_date_raw",
    "Day":                                                                                        "day_of_week",
    "Show announcement date:":                                                                    "announcement_date",
    "Email Address":                                                                              "contact_email",
    "What is the Bill Title for your gig or event?":                                              "act_name",
    "Promo bio (description of event or gig for website/socials):":                               "description",
    "Advanced ticket sale link:":                                                                 "ticket_link",
    "Is it OK for us to promote your gig/event on our website and/or socials?":                   "promo_ok",
    "Event start time?":                                                                          "start_time",
    "Door time?":                                                                                 "door_time",
    "Approx length of show/event:":                                                               "_length",
    "Are you charging for the gig/event? (If you are doing door/merch sales that require card payment, we advise that you bring your own card machine for taking electronic payments)": "ticketing",
    "Date tickets go on sale:":                                                                   "_ticket_sale_date",
    "Advanced ticket price:":                                                                     "ticket_price",
    "Day of ticket sale price:":                                                                  "_door_ticket_price",
    "Will you have a support act(s)?":                                                            "support_act",
    "Will you require a door person?":                                                            "_door_person_raw",
    "Will you be selling merch?":                                                                 "_merch",
    "There is a €150 fee for use of the room for gigs - this includes bar for the night and our sound engineer, regardless if you bring your own engineer.": "_fee_ack",
    "Do you have a photo(s) or an event/gig poster we can share on our website and social media?": "_poster_links",
    "Do you have links to other media we can share (music, socials, website)?":                   "media_links",
}


# ─── Helpers ────────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("xlsx", help="Path to Cobblestone form responses .xlsx")
    p.add_argument("--calendar-json", required=True,
                   help="Path to a JSON dump of calendar events (from MCP list_events).")
    p.add_argument("--confirm", action="store_true",
                   help="Actually write to DB (default: dry-run + report only).")
    p.add_argument("--report", default="import_review.md",
                   help="Markdown report path (default: import_review.md).")
    return p.parse_args()


def norm_email(e):
    if e is None:
        return ""
    return str(e).strip().lower()


def to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    # Try common formats — including the wonky "5/24/0026" the form sometimes captures
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S",
                "%-m/%-d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # Fallback: try the wonky 4-digit-year-with-leading-zero pattern (e.g. "5/24/0026")
    m = re.match(r"^(\d{1,2})/(\d{1,2})/0*(\d{2,4})$", s)
    if m:
        mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000
        try:
            return date(yy, mm, dd).isoformat()
        except ValueError:
            return None
    return None


def to_time_str(value):
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    return str(value).strip() or None


def door_person_required(raw):
    if not raw:
        return None  # not asked yet
    s = str(raw).strip().lower()
    if s.startswith("y") or "yes" in s:
        return 1
    if s.startswith("n") or "no" in s:
        return 0
    return None


# ─── Calendar match logic ────────────────────────────────────────────────────
def extract_cal_email(event):
    desc = (event.get("description") or "")
    m = EMAIL_RE.search(desc)
    return m.group(0).lower() if m else None


def cal_event_date(event):
    s = event.get("start")
    if isinstance(s, dict):
        d = s.get("date") or s.get("dateTime") or ""
    elif isinstance(s, str):
        d = s
    else:
        d = ""
    return d[:10] if d else None


def index_calendar(events):
    """Return dict: date_str -> list of events on that date."""
    by_date = defaultdict(list)
    for ev in events:
        d = cal_event_date(ev)
        if d:
            by_date[d].append(ev)
    return by_date


def _norm_title(s):
    """Normalize a title for fuzzy comparison: lowercase, strip non-alphanum, collapse spaces."""
    if not s:
        return ""
    s = re.sub(r"[^\w\s]", " ", str(s).lower())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _title_word_overlap(a, b):
    """Return fraction of significant-word overlap between two normalized titles."""
    stop = {"the", "a", "an", "and", "of", "in", "at", "on", "for", "to", "live"}
    wa = {w for w in a.split() if w not in stop and len(w) > 2}
    wb = {w for w in b.split() if w not in stop and len(w) > 2}
    if not wa or not wb:
        return 0.0
    overlap = len(wa & wb)
    return overlap / min(len(wa), len(wb))


def _is_title_match(form_title, cal_title):
    """Return True if titles look like the same act/event."""
    a = _norm_title(form_title)
    b = _norm_title(cal_title)
    if not a or not b:
        return False
    if a == b:
        return True
    # Either is a substring of the other (after normalization)
    if a in b or b in a:
        return True
    # Significant word overlap >= 60%
    return _title_word_overlap(a, b) >= 0.6


def match_form_to_calendar(form_row, cal_by_date, all_events):
    """Return (match_type, calendar_event_or_None).

    match_type ∈ {'exact', 'email_date_close', 'date_only', None}
    """
    form_email = form_row["contact_email"]
    form_date  = form_row["event_date"]
    form_title = form_row.get("act_name", "")

    if not form_date:
        return None, None

    on_date = cal_by_date.get(form_date, [])

    # 1a. Exact match: same date, email matches a calendar entry on that date
    for ev in on_date:
        if extract_cal_email(ev) == form_email:
            return "exact", ev

    # 1b. Same date, title matches (case-insensitive / fuzzy) — also exact
    for ev in on_date:
        if _is_title_match(form_title, ev.get("summary", "")):
            return "exact", ev

    # 2. Email match within ±N days (likely same booker rebooking nearby)
    if form_email:
        try:
            form_d = date.fromisoformat(form_date)
        except ValueError:
            return None, None
        for ev in all_events:
            if extract_cal_email(ev) != form_email:
                continue
            ev_date = cal_event_date(ev)
            if not ev_date:
                continue
            try:
                ev_d = date.fromisoformat(ev_date)
            except ValueError:
                continue
            if abs((ev_d - form_d).days) <= DATE_TOLERANCE_DAYS:
                return "email_date_close", ev

    # 3. Date match (no email match, no title match) — flag for review.
    #    Pick the calendar event with best title overlap for context.
    if on_date:
        best = max(on_date,
                   key=lambda ev: _title_word_overlap(_norm_title(form_title),
                                                      _norm_title(ev.get("summary", ""))))
        return "date_only", best

    return None, None


# ─── Form row parsing ───────────────────────────────────────────────────────
def build_form_row(headers, raw_row):
    out = {}
    for i, h in enumerate(headers):
        if h is None or h not in COL_MAP:
            continue
        key = COL_MAP[h]
        out[key] = raw_row[i] if i < len(raw_row) else None

    # Normalize fields for matching + DB
    out["contact_email"] = norm_email(out.get("contact_email"))
    out["event_date"]    = to_iso_date(out.get("_event_date_raw"))
    out["timestamp"]     = out.get("_timestamp")  # for dedupe
    out["start_time"]    = to_time_str(out.get("start_time"))
    out["door_time"]     = to_time_str(out.get("door_time"))
    out["door_fee_required"] = 1 if door_person_required(out.get("_door_person_raw")) == 1 else 0

    # ticket_price + support_act + media_links: stringify
    for k in ("ticket_price", "support_act", "media_links", "promo_ok",
              "ticketing", "announcement_date", "act_name", "description"):
        v = out.get(k)
        out[k] = str(v).strip() if v not in (None, "") else None

    # poster links go into media_links if media_links is empty
    poster = out.get("_poster_links")
    if poster and not out["media_links"]:
        out["media_links"] = str(poster).strip()
    elif poster:
        out["media_links"] = f"{out['media_links']}\nPoster/photo: {poster}"

    return out


def dedupe(rows):
    """Group by (email, event_date), keep the latest by timestamp."""
    by_key = {}
    duplicates = 0
    for r in rows:
        key = (r["contact_email"], r["event_date"])
        if not all(key):
            continue
        prev = by_key.get(key)
        if prev is None:
            by_key[key] = r
        else:
            # keep the more recent one
            if (r.get("timestamp") or datetime.min) > (prev.get("timestamp") or datetime.min):
                by_key[key] = r
            duplicates += 1
    return list(by_key.values()), duplicates


# ─── Markdown report ─────────────────────────────────────────────────────────
def fmt_row(r, cal_event=None, match_type=None):
    parts = []
    parts.append(f"- **{r.get('act_name') or '(no title)'}** — {r.get('event_date')} — `{r.get('contact_email')}`")
    if cal_event:
        parts.append(f"  - Calendar match ({match_type}): _{cal_event.get('summary','?')}_ on {cal_event_date(cal_event)}")
    extras = []
    if r.get("ticket_link"):    extras.append(f"ticket link: {r['ticket_link']}")
    if r.get("ticket_price"):   extras.append(f"price: {r['ticket_price']}")
    if r.get("support_act"):    extras.append(f"support: {r['support_act'][:60]}")
    if r.get("door_time"):      extras.append(f"doors: {r['door_time']}")
    if r.get("start_time"):     extras.append(f"start: {r['start_time']}")
    if r.get("door_fee_required"):
        extras.append("**door person requested**")
    if extras:
        parts.append("  - " + " · ".join(extras))
    return "\n".join(parts)


def write_report(path, auto_import, needs_review, pending, past_skipped, dup_count):
    lines = []
    lines.append("# Cobblestone form responses — import review")
    lines.append("")
    lines.append(f"_Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}_")
    lines.append("")
    lines.append("## Summary")
    lines.append(f"- ✅ Auto-import (status=confirmed): **{len(auto_import)}**")
    lines.append(f"- ⚠️ Needs review (status=inquiry, with best-guess match noted): **{len(needs_review)}**")
    lines.append(f"- 🟡 Pending (status=inquiry, no calendar match): **{len(pending)}**")
    lines.append(f"- ⚫ Past — skipped: **{len(past_skipped)}**")
    if dup_count:
        lines.append(f"- ↳ ({dup_count} duplicate submissions deduped, latest kept)")
    lines.append("")
    lines.append("To import, re-run with `--confirm`.")
    lines.append("To exclude specific rows, just remove them from the source xlsx (or tell me).")
    lines.append("")

    lines.append("## ✅ Auto-import")
    if not auto_import:
        lines.append("_(none)_")
    for r, ev in sorted(auto_import, key=lambda x: x[0]["event_date"]):
        lines.append(fmt_row(r, cal_event=ev, match_type="exact"))
    lines.append("")

    lines.append("## ⚠️ Needs review")
    if not needs_review:
        lines.append("_(none)_")
    for r, ev, mtype in sorted(needs_review, key=lambda x: x[0]["event_date"]):
        lines.append(fmt_row(r, cal_event=ev, match_type=mtype))
    lines.append("")

    lines.append("## 🟡 Pending (no calendar entry — likely never confirmed)")
    if not pending:
        lines.append("_(none)_")
    for r in sorted(pending, key=lambda x: x["event_date"]):
        lines.append(fmt_row(r))
    lines.append("")

    lines.append(f"## ⚫ Past — skipped ({len(past_skipped)} rows)")
    lines.append("_Not importing past-dated form submissions._")
    if past_skipped:
        lines.append("<details><summary>Show past rows</summary>")
        lines.append("")
        for r in sorted(past_skipped, key=lambda x: x["event_date"] or ""):
            lines.append(fmt_row(r))
        lines.append("")
        lines.append("</details>")
    lines.append("")

    with open(path, "w") as f:
        f.write("\n".join(lines))


# ─── DB writes (only with --confirm) ─────────────────────────────────────────
def to_booking_data(form_row, status, calendar_event_id=None, source_note="form-import"):
    """Build the dict shape expected by db.save_booking."""
    return {
        "venue":               "Backroom",
        "event_date":          form_row["event_date"],
        "day_of_week":         form_row.get("day_of_week"),
        "door_time":           form_row.get("door_time"),
        "start_time":          form_row.get("start_time"),
        "end_time":            None,
        "status":              status,
        "event_type":          "Gig",
        "act_name":            form_row.get("act_name") or "(no title)",
        "contact_name":        None,  # form doesn't capture separately from email
        "contact_email":       form_row["contact_email"],
        "contact_phone":       None,
        "expected_attendance": None,
        "description":         form_row.get("description"),
        "media_links":         form_row.get("media_links"),
        "ticketing":           form_row.get("ticketing"),
        "ticket_price":        form_row.get("ticket_price"),
        "ticket_link":         form_row.get("ticket_link"),
        "door_person":         None,
        "door_fee_required":   form_row.get("door_fee_required") or 0,
        "venue_fee_required":  1,
        "announcement_date":   form_row.get("announcement_date"),
        "support_act":         form_row.get("support_act"),
        "promo_ok":            form_row.get("promo_ok"),
        "notes":               source_note,
        "source":              "form-import",
    }


def _booking_already_imported(conn, act_name, event_date):
    """Check if a form-imported booking already exists for this act + date."""
    row = conn.execute(
        """SELECT id FROM bookings
           WHERE LOWER(TRIM(act_name)) = LOWER(TRIM(?))
             AND event_date = ?
             AND source = 'form-import'""",
        (act_name, event_date),
    ).fetchone()
    return row is not None


def insert_rows(auto_import, needs_review, pending):
    """Returns dict of counts. Idempotent — skips rows already imported."""
    db.init_db()
    conn = db.get_db()
    counts = {"confirmed": 0, "inquiry_review": 0, "inquiry_pending": 0,
              "skipped_existing": 0, "errors": 0}

    for r, ev in auto_import:
        try:
            if _booking_already_imported(conn, r.get("act_name") or "(no title)", r["event_date"]):
                counts["skipped_existing"] += 1
                continue
            data = to_booking_data(r, status="confirmed",
                                   source_note="form-import; matched calendar event exactly")
            bid = db.save_booking(data)
            if ev:
                ev_id = ev.get("id")
                if ev_id:
                    db.update_booking_field(bid, "google_calendar_event_id", ev_id, actor="import")
            db.add_booking_audit(bid, "import", "imported_from_form", "auto-import (exact match)")
            counts["confirmed"] += 1
        except Exception as e:
            print(f"  [auto] error on {r.get('act_name')}: {e}", file=sys.stderr)
            counts["errors"] += 1

    for r, ev, mtype in needs_review:
        try:
            if _booking_already_imported(conn, r.get("act_name") or "(no title)", r["event_date"]):
                counts["skipped_existing"] += 1
                continue
            note = f"form-import; needs review ({mtype} match — verify before confirming)"
            data = to_booking_data(r, status="inquiry", source_note=note)
            bid = db.save_booking(data)
            db.add_booking_audit(bid, "import", "imported_from_form_needs_review", note)
            counts["inquiry_review"] += 1
        except Exception as e:
            print(f"  [review] error on {r.get('act_name')}: {e}", file=sys.stderr)
            counts["errors"] += 1

    for r in pending:
        try:
            if _booking_already_imported(conn, r.get("act_name") or "(no title)", r["event_date"]):
                counts["skipped_existing"] += 1
                continue
            data = to_booking_data(r, status="inquiry",
                                   source_note="form-import; no calendar match — likely never confirmed")
            bid = db.save_booking(data)
            db.add_booking_audit(bid, "import", "imported_from_form_pending",
                                 "no calendar match")
            counts["inquiry_pending"] += 1
        except Exception as e:
            print(f"  [pending] error on {r.get('act_name')}: {e}", file=sys.stderr)
            counts["errors"] += 1

    return counts


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()
    today = date.today().isoformat()

    # Load form responses
    print(f"Loading form responses: {args.xlsx}")
    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb["Form Responses 1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"  {len(raw_rows)} rows in sheet")

    # Parse + dedupe
    parsed = [build_form_row(headers, r) for r in raw_rows]
    parsed = [r for r in parsed if r["event_date"] and r["contact_email"]]
    parsed, dup_count = dedupe(parsed)
    print(f"  {len(parsed)} unique (email, date) pairs after dedupe ({dup_count} duplicates)")

    # Load calendar
    print(f"Loading calendar dump: {args.calendar_json}")
    with open(args.calendar_json) as f:
        cal_data = json.load(f)
    cal_events = cal_data.get("events", cal_data) if isinstance(cal_data, dict) else cal_data
    print(f"  {len(cal_events)} calendar events")

    cal_by_date = index_calendar(cal_events)

    # Classify
    auto_import, needs_review, pending, past_skipped = [], [], [], []
    for r in parsed:
        if r["event_date"] < today:
            past_skipped.append(r)
            continue
        match_type, cal_event = match_form_to_calendar(r, cal_by_date, cal_events)
        if match_type == "exact":
            auto_import.append((r, cal_event))
        elif match_type in ("email_date_close", "date_only"):
            needs_review.append((r, cal_event, match_type))
        else:
            pending.append(r)

    print()
    print(f"Classification:")
    print(f"  auto-import      : {len(auto_import)}")
    print(f"  needs review     : {len(needs_review)}")
    print(f"  pending          : {len(pending)}")
    print(f"  past (skipped)   : {len(past_skipped)}")

    # Write report
    write_report(args.report, auto_import, needs_review, pending, past_skipped, dup_count)
    print(f"\nReport written to: {args.report}")

    if not args.confirm:
        print("\nDRY RUN — no DB changes. Re-run with --confirm to actually import.")
        return

    print("\nWriting to DB...")
    counts = insert_rows(auto_import, needs_review, pending)
    print(f"  status=confirmed     : {counts['confirmed']}")
    print(f"  status=inquiry (rev) : {counts['inquiry_review']}")
    print(f"  status=inquiry (pen) : {counts['inquiry_pending']}")
    print(f"  skipped (already in) : {counts['skipped_existing']}")
    print(f"  errors               : {counts['errors']}")


if __name__ == "__main__":
    main()
