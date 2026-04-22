"""One-shot import of historical bookings from Backroom_Booking_Tracker.xlsx.

Run once:
    python3 bookings_historical_import.py /path/to/Backroom_Booking_Tracker.xlsx [--dry-run]

Idempotent: skips rows whose (act_name, event_date) already exist with source='imported'.

Status mapping:
    Confirmed       -> confirmed
    Pending Info    -> tentative
    Inquiry         -> inquiry
    Date Unavailable, Cancelled -> cancelled
    Anything past-dated whose status was Confirmed -> completed
"""

import sys
from datetime import datetime, date, time
from openpyxl import load_workbook

import db


STATUS_MAP = {
    "confirmed":        "confirmed",
    "pending info":     "tentative",
    "tentative":        "tentative",
    "inquiry":          "inquiry",
    "date unavailable": "cancelled",
    "cancelled":        "cancelled",
    "canceled":         "cancelled",
    "completed":        "completed",
}


def _norm_status(value):
    if not value:
        return "inquiry"
    return STATUS_MAP.get(str(value).strip().lower(), "inquiry")


def _to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _to_time_str(value):
    if value is None:
        return None
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    return str(value).strip() or None


def _bool_paid(value):
    """Spreadsheet's 'Yes'/'No' or 'TRUE'/'FALSE' -> ISO timestamp or None."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ("yes", "y", "true", "1", "✓", "paid"):
        return datetime.now().isoformat()
    return None


def _bool_published(value):
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ("yes", "y", "true", "1", "scheduled", "✓", "published"):
        return datetime.now().isoformat()
    return None


def _door_person_norm(raw):
    if not raw:
        return None
    s = str(raw).strip().lower()
    if "pub" in s or "€50" in s:
        return "pub"
    if "own" in s:
        return "own"
    if "tbc" in s or "tba" in s:
        return "tbc"
    if "no" in s and "need" in s:
        return "none"
    return None


def _booking_already_imported(conn, act_name, event_date):
    row = conn.execute(
        """SELECT id FROM bookings
           WHERE act_name = ? AND event_date = ? AND source = 'imported'""",
        (act_name, event_date),
    ).fetchone()
    return row is not None


def import_main_sheet(ws, dry_run):
    """Bookings sheet — full schema."""
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    def cell(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    today_iso = date.today().isoformat()
    inserted = skipped = errored = 0
    conn = db.get_db()

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            act = cell(row, "Act / Event Name")
            evdate = _to_iso_date(cell(row, "Event Date"))
            if not act or not evdate:
                skipped += 1
                continue

            if _booking_already_imported(conn, str(act).strip(), evdate):
                skipped += 1
                continue

            status = _norm_status(cell(row, "Status"))
            # Past-dated Confirmed events become 'completed'
            if status == "confirmed" and evdate < today_iso:
                status = "completed"

            ev_type = (cell(row, "Event Type") or "Gig")
            door_person = _door_person_norm(cell(row, "Door Person"))
            venue_fee_required = 1 if (cell(row, "Venue Fee (€150)") not in (None, "", 0, "0", "no", "No")) else 1
            door_fee_required = 1 if door_person == "pub" else 0

            data = {
                "venue":               (cell(row, "Venue") or "Backroom"),
                "event_date":          evdate,
                "day_of_week":         cell(row, "Day"),
                "door_time":           _to_time_str(cell(row, "Door Time")),
                "start_time":          _to_time_str(cell(row, "Start Time")),
                "end_time":            None,
                "status":              status,
                "event_type":          str(ev_type).strip(),
                "act_name":            str(act).strip(),
                "contact_name":        cell(row, "Contact Name"),
                "contact_email":       cell(row, "Contact Email"),
                "contact_phone":       cell(row, "Contact Phone"),
                "expected_attendance": None,
                "description":         cell(row, "Description"),
                "media_links":         cell(row, "Media Links"),
                "ticketing":           cell(row, "Ticketing"),
                "ticket_price":        str(cell(row, "Adv. Ticket Price") or "") or None,
                "ticket_link":         cell(row, "Ticket Link"),
                "door_person":         door_person,
                "door_fee_required":   door_fee_required,
                "venue_fee_required":  venue_fee_required,
                "announcement_date":   str(cell(row, "Announcement Date") or "") or None,
                "support_act":         str(cell(row, "Support Act?") or "") or None,
                "promo_ok":            str(cell(row, "Promo OK?") or "") or None,
                "notes":               cell(row, "Notes"),
                "source":              "imported",
            }

            if dry_run:
                inserted += 1
                continue

            bid = db.save_booking(data)
            # Stamp paid timestamps + published timestamp from spreadsheet flags
            vfee_paid = _bool_paid(cell(row, "Venue Fee Paid"))
            dfee_paid = _bool_paid(cell(row, "Door Fee Paid"))
            published = _bool_published(cell(row, "On Website?"))
            if vfee_paid:
                db.update_booking_field(bid, "venue_fee_paid_at", vfee_paid, actor="import")
            if dfee_paid:
                db.update_booking_field(bid, "door_fee_paid_at", dfee_paid, actor="import")
            if published:
                db.update_booking_field(bid, "squarespace_published_at", published, actor="import")
            if cell(row, "Confirmation Sent?") and str(cell(row, "Confirmation Sent?")).strip().lower() in ("yes", "y", "true"):
                db.update_booking_field(bid, "confirmation_sent_at", datetime.now().isoformat(), actor="import")

            db.add_booking_audit(bid, "import", "imported_from_spreadsheet",
                                 f"Source row status={cell(row, 'Status')!r}")
            inserted += 1
        except Exception as e:
            errored += 1
            print(f"  [main] error on row '{cell(row, 'Act / Event Name')}': {e}")

    conn.close()
    return inserted, skipped, errored


def import_archive_sheet(ws, dry_run):
    """Archive (Pre-Apr 2026) — minimal columns, mark as completed/cancelled."""
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    def cell(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    today_iso = date.today().isoformat()
    inserted = skipped = errored = 0
    conn = db.get_db()

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            act = cell(row, "Act/Event")
            evdate = _to_iso_date(cell(row, "Event Date"))
            if not act or not evdate:
                skipped += 1
                continue

            if _booking_already_imported(conn, str(act).strip(), evdate):
                skipped += 1
                continue

            old_status = _norm_status(cell(row, "Old Status"))
            if old_status == "cancelled":
                status = "cancelled"
            elif evdate < today_iso:
                status = "completed"
            else:
                status = "tentative"

            data = {
                "venue":               "Backroom",
                "event_date":          evdate,
                "day_of_week":         None,
                "door_time":           None,
                "start_time":          None,
                "end_time":            None,
                "status":              status,
                "event_type":          "Gig",
                "act_name":            str(act).strip(),
                "contact_name":        None,
                "contact_email":       cell(row, "Email"),
                "contact_phone":       None,
                "expected_attendance": None,
                "description":         cell(row, "Description"),
                "media_links":         None,
                "ticketing":           cell(row, "Charging"),
                "ticket_price":        None,
                "ticket_link":         None,
                "door_person":         _door_person_norm(cell(row, "Door Person")),
                "door_fee_required":   0,
                "venue_fee_required":  1,
                "announcement_date":   None,
                "support_act":         None,
                "promo_ok":            None,
                "notes":               f"Imported from archive sheet (Pre-Apr 2026). Original status: {cell(row, 'Old Status')}",
                "source":              "imported",
            }

            if dry_run:
                inserted += 1
                continue

            bid = db.save_booking(data)
            db.add_booking_audit(bid, "import", "imported_from_archive_sheet",
                                 f"Original status: {cell(row, 'Old Status')!r}")
            inserted += 1
        except Exception as e:
            errored += 1
            print(f"  [archive] error on row '{cell(row, 'Act/Event')}': {e}")

    conn.close()
    return inserted, skipped, errored


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    xlsx_path = sys.argv[1]
    dry_run = "--dry-run" in sys.argv[2:]

    print(f"Loading {xlsx_path}{' (DRY RUN)' if dry_run else ''}")
    db.init_db()
    wb = load_workbook(xlsx_path, data_only=True)

    if "Bookings" in wb.sheetnames:
        print("Importing 'Bookings' sheet...")
        ins, skip, err = import_main_sheet(wb["Bookings"], dry_run)
        print(f"  inserted={ins} skipped={skip} errors={err}")

    archive_name = next((n for n in wb.sheetnames if n.startswith("Archive")), None)
    if archive_name:
        print(f"Importing '{archive_name}' sheet...")
        ins, skip, err = import_archive_sheet(wb[archive_name], dry_run)
        print(f"  inserted={ins} skipped={skip} errors={err}")

    counts = db.booking_counts()
    print(f"\nDone. Current counts: {counts}")


if __name__ == "__main__":
    main()
